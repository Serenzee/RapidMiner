import subprocess
import sys

# При запуске этого скрипта автоматически запускается results_bot.py как отдельный процесс.
if __name__ == "__main__":
    subprocess.Popen([sys.executable, "results_bot.py"])

import asyncio
import random
import sqlite3
import pandas as pd
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from questions_pr1 import QUESTIONS as QUESTIONS_PR1
from questions_pr2 import QUESTIONS as QUESTIONS_PR2
from questions_pr3 import QUESTIONS as QUESTIONS_PR3
from questions_pr4 import QUESTIONS as QUESTIONS_PR4
from questions_pr5 import QUESTIONS as QUESTIONS_PR5
from questions_pr6 import QUESTIONS as QUESTIONS_PR6
from questions_pr7 import QUESTIONS as QUESTIONS_PR7
from questions_pr8 import QUESTIONS as QUESTIONS_PR8
from questions_pr9 import QUESTIONS as QUESTIONS_PR9
from questions_pr10 import QUESTIONS as QUESTIONS_PR10
from questions_pr11 import QUESTIONS as QUESTIONS_PR11
from questions_pr12 import QUESTIONS as QUESTIONS_PR12
from questions_pr13 import QUESTIONS as QUESTIONS_PR13
from questions_pr14 import QUESTIONS as QUESTIONS_PR14

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

API_TOKEN = 'API_TG_BOT'

# Список всех тестов/практик
PRACTICES = [
    "Знакомство с RapidMiner",
    "Кластерный анализ",
    "Визуализации в RapidMiner",
    "Ассоциативные правила",
    "Решающие деревья",
    "Изучение возможностей ML моделей",
    "Изучение Deep Learning в RapidMiner",
    "Анализ текста (Text Mining)",
    "Web Scraping и анализ данных",
    "Расчет географических расстояний",
    "RandomForest и линейная регрессия",
    "Прогнозирование временных рядов",
    "Интерпретация моделей",
    "Аномалий в IoT-данных"
]

# Массив с вопросами по тестам
QUESTIONS_BY_PRACTICE = [
    QUESTIONS_PR1, QUESTIONS_PR2, QUESTIONS_PR3, QUESTIONS_PR4, QUESTIONS_PR5, QUESTIONS_PR6, QUESTIONS_PR7,
    QUESTIONS_PR8, QUESTIONS_PR9, QUESTIONS_PR10, QUESTIONS_PR11, QUESTIONS_PR12, QUESTIONS_PR13, QUESTIONS_PR14
]

bot = Bot(token=API_TOKEN)
dp = Dispatcher()
user_state = {}          # Текущее состояние пользователя (для FSM)
registration_state = {}  # Состояние регистрации

def init_db():
    """Создание всех необходимых таблиц, если их нет."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT,
            group_number TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS progress (
            user_id INTEGER PRIMARY KEY,
            current_practice INTEGER
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS results (
            user_id INTEGER,
            practice_idx INTEGER,
            score INTEGER,
            PRIMARY KEY (user_id, practice_idx)
        )
    ''')
    conn.commit()
    conn.close()

def user_exists(user_id):
    """Проверяет, зарегистрирован ли пользователь по user_id."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE user_id = ?', (user_id,))
    result = c.fetchone()
    conn.close()
    return result is not None

def full_name_exists(full_name):
    """Проверяет, есть ли уже такое ФИО (чтобы нельзя было дублировать)."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE full_name = ?', (full_name,))
    result = c.fetchone()
    conn.close()
    return result is not None

def save_user(user_id, full_name, group_number):
    """Сохраняет пользователя и его прогресс. После этого обновляет таблицу Excel."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('INSERT OR REPLACE INTO users (user_id, full_name, group_number) VALUES (?, ?, ?)',
              (user_id, full_name, group_number))
    c.execute('INSERT OR IGNORE INTO progress (user_id, current_practice) VALUES (?, 0)', (user_id,))
    conn.commit()
    conn.close()
    generate_results_excel()

def get_current_practice(user_id):
    """Возвращает индекс текущего теста (по факту последний завершённый)."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('SELECT current_practice FROM progress WHERE user_id = ?', (user_id,))
    result = c.fetchone()
    conn.close()
    return result[0] if result else 0

def advance_practice(user_id):
    """Увеличивает номер текущей доступной практики и обновляет Excel."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('SELECT current_practice FROM progress WHERE user_id = ?', (user_id,))
    res = c.fetchone()
    if res:
        c.execute('UPDATE progress SET current_practice = current_practice + 1 WHERE user_id = ?', (user_id,))
    else:
        c.execute('INSERT INTO progress (user_id, current_practice) VALUES (?, 1)', (user_id,))
    conn.commit()
    conn.close()
    generate_results_excel()

def save_result(user_id, practice_idx, score):
    """Сохраняет результат прохождения теста и обновляет Excel."""
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute('INSERT OR REPLACE INTO results (user_id, practice_idx, score) VALUES (?, ?, ?)', (user_id, practice_idx, score))
    conn.commit()
    conn.close()
    generate_results_excel()

async def send_long_message(message, text, reply_markup=None):
    """Отправка длинных сообщений по частям (ограничение Telegram)."""
    MAX_LEN = 4096
    for i in range(0, len(text), MAX_LEN):
        await message.answer(text[i:i+MAX_LEN], reply_markup=reply_markup if i + MAX_LEN >= len(text) else None)

def make_practice_menu_single(user_id):
    """Генерирует меню с только одним доступным тестом для пользователя."""
    current = get_current_practice(user_id)
    kb = [
        [types.InlineKeyboardButton(text=f"{current+1}. {PRACTICES[current]}", callback_data=f"select_practice_{current}")]
    ]
    text = f"Доступен тест {current+1}.\nДля доступа к следующему завершите текущий."
    return text, types.InlineKeyboardMarkup(inline_keyboard=kb)

def format_full_name(s: str):
    """Проверяет корректность ФИО (3 слова), каждое с заглавной буквы. Возвращает отформатированную строку или None."""
    parts = s.strip().split()
    if len(parts) != 3:
        return None
    return ' '.join(w.capitalize() for w in parts)

def check_group(s: str):
    """Проверяет корректность группы (одно слово, латиница/кириллица + цифры, преобразует к капсам)."""
    s = s.strip().upper()
    import re
    return bool(re.fullmatch(r'[A-ZА-ЯЁ]+\d+', s)), s

@dp.message(Command("start"))
async def process_start(message: types.Message):
    """Обработчик команды /start. Запускает регистрацию или показывает меню выбора теста."""
    user_id = message.from_user.id
    if not user_exists(user_id):
        registration_state[user_id] = {"step": "wait_fullname"}
        await message.answer("Здравствуйте! Для прохождения теста введите, пожалуйста, ФИО (Фамилия Имя Отчество):")
    else:
        text, kb = make_practice_menu_single(user_id)
        await message.answer(text, reply_markup=kb)

@dp.message()
async def registration_handler(message: types.Message):
    """FSM обработчик регистрации пользователя."""
    user_id = message.from_user.id
    if user_id not in registration_state:
        if not user_exists(user_id):
            await message.answer("Пожалуйста, сначала зарегистрируйтесь через /start.")
        return

    state = registration_state[user_id]
    if state["step"] == "wait_fullname":
        name_input = message.text.strip()
        formatted_name = format_full_name(name_input)
        if not formatted_name:
            await message.answer("Ошибка: ФИО должно содержать три слова (Фамилия Имя Отчество), каждое с большой буквы. Попробуйте снова.")
            return
        if full_name_exists(formatted_name):
            await message.answer("Ошибка: Пользователь с таким ФИО уже зарегистрирован. Введите другое ФИО.")
            return
        registration_state[user_id]["full_name"] = formatted_name
        registration_state[user_id]["step"] = "wait_group"
        await message.answer("Теперь введите номер группы (например, БИБ251):")
    elif state["step"] == "wait_group":
        group = message.text.strip()
        is_ok, formatted_group = check_group(group)
        if not is_ok:
            await message.answer("Ошибка: Номер группы должен быть одним словом на латинице и цифрах, например, БИБ215. Попробуйте снова.")
            return
        full_name = registration_state[user_id].get("full_name", "")
        save_user(user_id, full_name, formatted_group)
        del registration_state[user_id]
        await message.answer("Спасибо, регистрация завершена! Теперь вы можете пройти тестирование.")
        text, kb = make_practice_menu_single(user_id)
        await message.answer(text, reply_markup=kb)

def option_buttons(practice_idx, active_msg_id):
    """Генерирует кнопки выбора ответа для теста."""
    keyboard = [
        [
            types.InlineKeyboardButton(text="A", callback_data=f"pick_{practice_idx}_0_{active_msg_id}"),
            types.InlineKeyboardButton(text="B", callback_data=f"pick_{practice_idx}_1_{active_msg_id}"),
            types.InlineKeyboardButton(text="C", callback_data=f"pick_{practice_idx}_2_{active_msg_id}"),
            types.InlineKeyboardButton(text="D", callback_data=f"pick_{practice_idx}_3_{active_msg_id}"),
        ]
    ]
    return types.InlineKeyboardMarkup(inline_keyboard=keyboard)

async def send_question(message, user_id):
    """Отправляет пользователю очередной вопрос теста или результат после завершения."""
    if not user_exists(user_id):
        await message.answer("Пожалуйста, сначала зарегистрируйтесь через /start.")
        return

    state = user_state[user_id]
    idx = state["current"]
    questions = state["questions"]
    practice_idx = state["practice"]
    total_questions = state["total_questions"]

    if idx >= total_questions:
        correct = state["correct"]
        save_result(user_id, practice_idx, correct)
        result = f"Тест завершён!\n\nПравильных ответов: {correct}/{total_questions}\n"
        kb = types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(text="Меню", callback_data="main_menu")]
        ])
        await send_long_message(message, result, reply_markup=kb)
        user_state[user_id]["active_msg_id"] = None
        user_state[user_id]["active_test"] = False
    else:
        q_idx = state["order"][idx]
        q = questions[q_idx]
        option_indices = list(range(len(q["options"])))
        random.shuffle(option_indices)
        shuffled_options = [q["options"][i] for i in option_indices]
        correct_shuffled = option_indices.index(q["answer"])
        state["shuffle"] = option_indices
        state["correct_shuffled"] = correct_shuffled

        text = f"Вопрос {idx+1} из {total_questions}:\n\n{q['question']}\n\n" + \
            "\n".join([f"{chr(65+i)}) {opt}" for i, opt in enumerate(shuffled_options)])
        sent = await message.answer(text)
        user_state[user_id]["active_msg_id"] = sent.message_id
        user_state[user_id]["active_test"] = True
        await bot.edit_message_reply_markup(chat_id=sent.chat.id, message_id=sent.message_id,
                                            reply_markup=option_buttons(practice_idx, sent.message_id))

@dp.callback_query(lambda c: c.data == "denied")
async def denied_callback(callback: types.CallbackQuery):
    """Ответ на попытку выбрать недоступный тест."""
    await callback.answer("Доступ запрещен", show_alert=True)

@dp.callback_query(lambda c: c.data.startswith("select_practice_"))
async def select_practice(callback: types.CallbackQuery):
    """Обработчик выбора теста из меню."""
    user_id = callback.from_user.id
    if not user_exists(user_id):
        await callback.message.answer("Пожалуйста, сначала зарегистрируйтесь через /start.")
        await callback.answer()
        return

    idx = int(callback.data.split("_")[-1])
    allowed = get_current_practice(user_id)
    state = user_state.get(user_id, {})
    if idx != allowed:
        await callback.answer("Доступ запрещен", show_alert=True)
        return
    if state.get("active_test"):
        await callback.answer("Вы уже проходите этот тест. Сначала завершите текущий.", show_alert=True)
        return
    if idx < len(QUESTIONS_BY_PRACTICE) and QUESTIONS_BY_PRACTICE[idx]:
        questions = QUESTIONS_BY_PRACTICE[idx]
        if len(questions) < 10:
            await callback.message.answer("Для этой практики недостаточно вопросов для теста.")
            await callback.answer()
            return
        order = random.sample(range(len(questions)), 10)
        user_state[user_id] = {
            "practice": idx,
            "order": order,
            "current": 0,
            "correct": 0,
            "questions": questions,
            "total_questions": 10,
            "active_msg_id": None,
            "active_test": True
        }
        await send_question(callback.message, user_id)
    else:
        await callback.message.answer(f"Пока для этой практики нет теста.\nОжидайте обновлений!")
    await callback.answer()

@dp.callback_query(lambda c: c.data.startswith("pick_"))
async def handle_pick(callback: types.CallbackQuery):
    """Обработчик выбора варианта ответа в тесте."""
    user_id = callback.from_user.id
    if not user_exists(user_id):
        await callback.message.answer("Пожалуйста, сначала зарегистрируйтесь через /start.")
        await callback.answer()
        return

    if user_id not in user_state:
        await callback.message.answer("Сначала выберите практику через /start.")
        await callback.answer()
        return

    try:
        _, practice_idx, picked, msg_id = callback.data.split("_")
    except Exception:
        await callback.answer("Доступ запрещен", show_alert=True)
        return

    practice_idx = int(practice_idx)
    picked = int(picked)
    msg_id = int(msg_id)
    state = user_state[user_id]

    # Только последняя активная кнопка валидна
    if state.get("active_msg_id") != callback.message.message_id:
        await callback.answer("Доступ запрещен", show_alert=True)
        return

    idx = state["current"]
    questions = state["questions"]
    q_idx = state["order"][idx]
    q = questions[q_idx]
    option_indices = state.get("shuffle", list(range(4)))
    correct_shuffled = state.get("correct_shuffled", q["answer"])

    if picked == correct_shuffled:
        state["correct"] += 1
    state["current"] += 1

    if state["current"] >= state["total_questions"]:
        if get_current_practice(user_id) == state["practice"]:
            advance_practice(user_id)
    await send_question(callback.message, user_id)
    await callback.answer()

@dp.callback_query(lambda c: c.data == "main_menu")
async def to_main_menu(callback: types.CallbackQuery):
    """Возвращает пользователя в главное меню выбора теста."""
    user_id = callback.from_user.id
    if not user_exists(user_id):
        await callback.message.answer("Пожалуйста, сначала зарегистрируйтесь через /start.")
        await callback.answer()
        return

    text, kb = make_practice_menu_single(user_id)
    await callback.message.answer(text, reply_markup=kb)
    await callback.answer()

def generate_results_excel():
    """
    Формирует Excel-файл с результатами по всем группам.
    Страницы - это отдельные группы, автоширина и границы у таблицы.
    В каждой группе ФИО идут по алфавиту.
    В заголовках теста — номер и название, через перенос строки.
    """
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT user_id, full_name, group_number FROM users")
    users = {row[0]: {'full_name': row[1], 'group_number': row[2]} for row in c.fetchall()}
    c.execute("SELECT user_id, practice_idx, score FROM results")
    results = c.fetchall()
    conn.close()

    if not users or not results:
        return

    data_by_group = {}
    for user_id, user in users.items():
        group = user['group_number']
        if group not in data_by_group:
            data_by_group[group] = []
        scores = [""] * len(PRACTICES)
        for res in results:
            if res[0] == user_id:
                scores[res[1]] = f"{res[2]}/10"
        data_by_group[group].append([user['full_name']] + scores)

    with pd.ExcelWriter("results.xlsx", engine="openpyxl") as writer:
        # Заголовок с переносом строки между номером и названием теста
        columns = ["ФИО"] + [f"Тест {i+1}.\n{PRACTICES[i]}" for i in range(len(PRACTICES))]
        for group, students in data_by_group.items():
            # Сортировка студентов внутри группы по ФИО
            students.sort(key=lambda x: x[0])
            df = pd.DataFrame(students, columns=columns)
            df.to_excel(writer, sheet_name=str(group), index=False)

    # Форматирование: автоширина и тонкая граница для всех ячеек
    wb = load_workbook("results.xlsx")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        # автоширина ФИО, фикс ширина тестов
        for col_idx, col in enumerate(ws.columns, 1):
            column = col[0].column_letter
            if col_idx == 1:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[column].width = max_length + 2
            else:
                ws.column_dimensions[column].width = 18
        # границы
        for i in range(1, max_row+1):
            for j in range(1, max_col+1):
                ws.cell(row=i, column=j).border = thin_border
        # центровка и перенос заголовков
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # центровка баллов
        for i in range(2, max_row+1):
            for j in range(2, max_col+1):
                ws.cell(row=i, column=j).alignment = Alignment(horizontal="center", vertical="center")
    wb.save("results.xlsx")


async def on_startup():
    """Выполняется при запуске программы. Создает БД при первом запуске."""
    init_db()
    # generate_results_excel()  # Не генерируем файл если нет данных

async def main():
    """Главная асинхронная функция — запуск Telegram-бота."""
    await on_startup()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
